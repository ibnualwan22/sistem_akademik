# core/admin.py - VERSI FINAL YANG RAPI DAN BENAR

from django.contrib import admin
from django.contrib.admin.models import LogEntry
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .models import Fan, SKS, Santri, RiwayatTes, Pengurus
from datetime import date

# ==========================================================================
# FUNGSI UNTUK ADMIN ACTION (didefinisikan di luar kelas)
# ==========================================================================
def export_as_excel(modeladmin, request, queryset):
    """
    Aksi admin untuk mengekspor data RiwayatTes yang dipilih ke format Excel,
    dikelompokkan per penguji.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="rekap_tes_{date.today().isoformat()}.xlsx"'
    workbook = Workbook()

    tes_by_examiner = {}
    for tes in queryset.order_by('penguji__nama_lengkap', 'santri__nama_lengkap'):
        penguji_nama = tes.penguji.nama_lengkap if tes.penguji else "Tanpa Penguji"
        if penguji_nama not in tes_by_examiner:
            tes_by_examiner[penguji_nama] = []
        tes_by_examiner[penguji_nama].append(tes)

    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    for penguji, tes_list in tes_by_examiner.items():
        worksheet = workbook.create_sheet(title=penguji[:31])
        worksheet.merge_cells('A1:E1')
        worksheet['A1'] = 'REKAP PENDAFTARAN TES ASRAMA TAKHOSSUS'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A1'].alignment = center_align
        worksheet.merge_cells('A2:E2')
        worksheet['A2'] = f"PENGUJI : {penguji.upper()}"
        worksheet['A2'].font = header_font
        if tes_list:
            tanggal_tes = tes_list[0].tanggal_pelaksanaan.strftime('%d %B %Y')
            worksheet.merge_cells('A3:E3')
            worksheet['A3'] = f"Tanggal : {tanggal_tes}"
        
        headers = ['NO', 'NAMA PESERTA', 'KITAB', 'NILAI', 'KETERANGAN']
        worksheet.append(headers)
        for cell in worksheet[5]:
            cell.font = header_font
            cell.alignment = center_align

        for i, tes in enumerate(tes_list, 1):
            keterangan = "-"
            if tes.status_tes == 'Terdaftar': keterangan = 'Baru Mendaftar'
            elif tes.nilai is not None: keterangan = tes.status_kelulusan
            row_data = [i, tes.santri.nama_lengkap, tes.sks.nama_sks, tes.nilai if tes.nilai is not None else '', keterangan]
            worksheet.append(row_data)
            worksheet.cell(row=worksheet.max_row, column=1).alignment = Alignment(horizontal='center')
        
        for col, width in [('A', 5), ('B', 35), ('C', 35), ('D', 10), ('E', 20)]:
            worksheet.column_dimensions[col].width = width
            
    workbook.save(response)
    return response

export_as_excel.short_description = "Ekspor Data Terpilih ke Excel (per Penguji)"


# ==========================================================================
# KELAS-KELAS ADMIN
# ==========================================================================

class StatusKelulusanFilter(admin.SimpleListFilter):
    title = 'Status Kelulusan'
    parameter_name = 'status_kelulusan'
    def lookups(self, request, model_admin): return (('lulus', 'Lulus'), ('mengulang', 'Mengulang'),)
    def queryset(self, request, queryset):
        if self.value() == 'lulus': return queryset.filter(id__in=[tes.id for tes in queryset if tes.status_kelulusan == 'Lulus'])
        if self.value() == 'mengulang': return queryset.filter(id__in=[tes.id for tes in queryset if tes.status_kelulusan == 'Mengulang'])

@admin.register(Fan)
class FanAdmin(admin.ModelAdmin):
    list_display = ('nama_fan', 'urutan', 'target_durasi_hari')
    search_fields = ('nama_fan',)
    ordering = ('urutan',)

@admin.register(SKS)
class SKSAdmin(admin.ModelAdmin):
    list_display = ('nama_sks', 'fan', 'nilai_minimal')
    list_filter = ('fan',)
    search_fields = ('nama_sks',)
    autocomplete_fields = ('fan',)

@admin.register(Santri)
class SantriAdmin(admin.ModelAdmin):
    list_display = ('nama_lengkap', 'id_santri', 'status')
    list_filter = ('status',)
    search_fields = ('nama_lengkap', 'id_santri')
    list_per_page = 20

@admin.register(RiwayatTes)
class RiwayatTesAdmin(admin.ModelAdmin):
    list_display = ('get_nama_santri', 'get_nama_sks', 'tanggal_pelaksanaan', 'penguji', 'nilai', 'status_tes', 'status_kelulusan')
    list_filter = ('status_tes', 'tanggal_pelaksanaan', 'sks__fan', 'penguji', StatusKelulusanFilter)
    search_fields = ('santri__nama_lengkap', 'sks__nama_sks', 'penguji__nama_lengkap')
    autocomplete_fields = ('santri', 'sks', 'penguji')
    list_per_page = 25
    fieldsets = (
        ('Informasi Pendaftaran', {'fields': ('santri', 'sks', 'penguji', 'tanggal_pelaksanaan')}),
        ('Status & Penilaian', {'fields': ('status_tes', 'nilai', 'tanggal_pendaftaran')}),
    )
    actions = [export_as_excel]

    # --- METHOD-METHOD INI HARUS BERADA DI DALAM KELAS ---
    def get_changeform_initial_data(self, request):
        return {'tanggal_pendaftaran': date.today()}

    @admin.display(description='Nama Santri', ordering='santri__nama_lengkap')
    def get_nama_santri(self, obj):
        return obj.santri.nama_lengkap

    @admin.display(description='SKS / Kitab', ordering='sks__nama_sks')
    def get_nama_sks(self, obj):
        return obj.sks.nama_sks

@admin.register(Pengurus)
class PengurusAdmin(admin.ModelAdmin):
    list_display = ('nama_lengkap', 'jabatan', 'nomor_whatsapp')
    search_fields = ('nama_lengkap', 'jabatan')

@admin.register(LogEntry)
class LogEntryAdmin(admin.ModelAdmin):
    readonly_fields = ('action_time', 'user', 'content_type', 'object_id', 'object_repr', 'action_flag', 'change_message')
    list_display = ('action_time', 'user', 'get_action_description', 'object_repr', 'get_content_type')
    list_filter = ('action_time', 'user', 'content_type')
    def has_add_permission(self, request): return False
    def has_change_permission(self, request, obj=None): return False
    def has_delete_permission(self, request, obj=None): return False
    @admin.display(description='Objek')
    def get_content_type(self, obj): return obj.content_type.name
    @admin.display(description='Aksi', ordering='action_flag')
    def get_action_description(self, obj): return {1: "➕ Penambahan", 2: "✏️ Perubahan", 3: "❌ Penghapusan"}.get(obj.action_flag, '-')